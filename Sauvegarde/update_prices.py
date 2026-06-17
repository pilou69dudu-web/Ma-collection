import os
import sys
import subprocess
import shutil
import re
import time

def installer_si_manquant(module_name):
    try:
        __import__(module_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", module_name])

installer_si_manquant("openpyxl")
installer_si_manquant("requests")

import openpyxl
import requests

# =====================================================================
# CONFIGURATION OFFICIELLE
# =====================================================================
DISCOGS_TOKEN = "MjiPBNhclfyzkraPghTWfzHbjaqKKMzLYmCeoBjR"
FICHIER_ORIGINAL = "00_Mes vinyles.xlsx"
FICHIER_TMP = "00_Mes vinyles_TEMP.xlsx"
# =====================================================================

def extraire_id_discogs(url):
    if not url:
        return None
    match = re.search(r"release/(\d+)", str(url))
    return match.group(1) if match else None

def recuperer_vrai_prix_low(release_id):
    """Va chercher directement le prix plancher (Low) de l'historique d'achat"""
    url = f"https://api.discogs.com/releases/{release_id}"
    headers = {
        "User-Agent": "VinylothequeAutomatique/6.0 (Windows NT 10.0; Win64; x64)",
        "Authorization": f"Discogs token={DISCOGS_TOKEN}"
    }
    
    try:
        reponse = requests.get(url, headers=headers, timeout=12)
        
        # Sécurité si le serveur sature
        if reponse.status_code == 429:
            time.sleep(15)
            reponse = requests.get(url, headers=headers, timeout=12)

        if reponse.status_code == 200:
            data = reponse.json()
            stats = data.get("marketplace_stats", {})
            
            if stats:
                # On extrait la valeur brute 'lowest' (Prix Bas) constatée sur le marché
                if "lowest" in stats and stats["lowest"].get("value"):
                    valeur_usd = float(stats["lowest"]["value"])
                    # Conversion brute en euros (0.89) sans AUCUN coefficient ajouté
                    prix_euros = round(valeur_usd * 0.89, 2)
                    return f"{prix_euros}€"
                    
                # Secours si le prix bas n'est pas renseigné, on prend le médian
                if "median" in stats and stats["median"].get("value"):
                    valeur_usd = float(stats["median"]["value"])
                    return f"{round(valeur_usd * 0.89, 2)}€"
                    
    except Exception:
        pass
    return None

def main():
    print("=====================================================================")
    print("      MISE À JOUR AUTONOME ET SÉCURISÉE DES PRIX (PRIX BAS 'LOW')     ")
    print("=====================================================================")

    if not os.path.exists(FICHIER_ORIGINAL):
        print(f"❌ Impossible de trouver le fichier : {FICHIER_ORIGINAL}")
        return

    # Gestion propre du fichier temporaire bloqué
    if os.path.exists(FICHIER_TMP):
        try:
            os.remove(FICHIER_TMP)
        except Exception:
            print("❌ Le fichier temporaire est verrouillé. Assurez-vous d'avoir FERMÉ EXCEL.")
            return

    try:
        shutil.copyfile(FICHIER_ORIGINAL, FICHIER_TMP)
    except PermissionError:
        print("❌ Permission refusée. Veuillez fermer votre fichier Excel avant de lancer.")
        return

    try:
        wb = openpyxl.load_workbook(FICHIER_TMP)
        ws = wb.active
    except Exception as e:
        print(f"❌ Erreur d'ouverture Excel : {e}")
        return

    # Détection automatique des colonnes
    header_row = 1
    for r in range(1, 4):
        cells = [str(ws.cell(row=r, column=c).value).lower() for c in range(1, ws.max_column + 1)]
        if "artiste" in cells or "lien" in cells:
            header_row = r
            break

    headers = {str(ws.cell(row=header_row, column=c).value).strip(): c for c in range(1, ws.max_column + 1)}
    col_lien = headers.get('Lien')
    col_artiste = headers.get('ARTISTE')
    col_prix = 18  # Colonne R (Prix Haut / Prix Bas)

    print("🚀 Le script s'exécute en arrière-plan. Vous pouvez partir de l'ordinateur.")
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Évite d'écraser si un prix est déjà récupéré lors d'une précédente session
        valeur_actuelle = ws.cell(row=row_idx, column=col_prix).value
        if valeur_actuelle and "€" in str(valeur_actuelle):
            continue

        artiste = ws.cell(row=row_idx, column=col_artiste).value if col_artiste else f"Ligne {row_idx}"
        lien = ws.cell(row=row_idx, column=col_lien).value
        
        id_discogs = extraire_id_discogs(lien)
        if not id_discogs:
            continue

        print(f"🔍 [{row_idx - header_row}/{ws.max_row - header_row}] {artiste}... ", end="", flush=True)
        
        vrai_prix = recuperer_vrai_prix_low(id_discogs)

        if vrai_prix:
            ws.cell(row=row_idx, column=col_prix).value = vrai_prix
            print(f"✅ VALEUR ENREGISTRÉE : {vrai_prix}")
        else:
            ws.cell(row=row_idx, column=col_prix).value = "Non vendu"
            print("💾 (Pas de cote brute)")

        # Sauvegarde automatique ligne par ligne pour sécuriser les données
        try:
            wb.save(FICHIER_TMP)
        except PermissionError:
            print("\n❌ Excel a été ouvert pendant l'exécution. Arrêt de sécurité.")
            wb.close()
            return
            
        # Temporisation de sécurité pour respecter les quotas de Discogs
        time.sleep(3.0)

    wb.close()
    try:
        shutil.move(FICHIER_TMP, FICHIER_ORIGINAL)
        print("\n✅ Félicitations ! Votre Vinylthèque est entièrement à jour.")
    except Exception:
        print("\n⚠️ Importation finale impossible. Fermez Excel pour appliquer les changements.")

if __name__ == "__main__":
    main()