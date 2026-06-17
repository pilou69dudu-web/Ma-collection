import os
import openpyxl
import json

# Chemins des fichiers
EXCEL_VINYLES = "00_Mes vinyles.xlsx"
EXCEL_ACHAT = "01_Liste achat.xlsx"
INDEX_HTML = "index.html"
WANTED_HTML = "Wanted.html"

print("=========================================================================")
print("               GÉNÉRATION DE LA PAGE WEB ET ENVOI SUR GITHUB             ")
print("=========================================================================")
print("\n1. Génération de la collection de Vinyles...")

# ---- CHARGEMENT DE LA COLLECTION PRINCIPALE ----
if not os.path.exists(EXCEL_VINYLES):
    print(f"Erreur : Le fichier {EXCEL_VINYLES} est introuvable.")
    exit()

wb_v = openpyxl.load_workbook(EXCEL_VINYLES, data_only=True)
sheet_v = wb_v.active

# Récupération du compteur global affiché en A1 (ex: 3741)
compteur_global = sheet_v["A1"].value if sheet_v["A1"].value else 0
print(f"🎯 Compteur récupéré en A1 : {compteur_global} disques au total.")

liste_vinyles = []
# On parcourt de la ligne 2 à 3192 (vos 3191 lignes de données)
for row_idx in range(2, 3193):
    type_album = sheet_v.cell(row=row_idx, column=3).value     # Col C: Album/Compils/Single
    annee = sheet_v.cell(row=row_idx, column=4).value          # Col D: Année
    quantite = sheet_v.cell(row=row_idx, column=5).value       # Col E: Qté
    pays = sheet_v.cell(row=row_idx, column=6).value           # Col F: Pays
    genre = sheet_v.cell(row=row_idx, column=7).value          # Col G: Genre
    commentaire = sheet_v.cell(row=row_idx, column=8).value    # Col H: Picture / Commentaire / Prix
    artiste = sheet_v.cell(row=row_idx, column=9).value        # Col I: ARTISTE
    titre_face_a = sheet_v.cell(row=row_idx, column=11).value   # Col K: Titre Face A
    titre_face_b = sheet_v.cell(row=row_idx, column=13).value   # Col M: Titre Face B

    if not artiste and not genre:
        continue

    # Construction de l'objet vinyle (nettoyé et sécurisé pour le JavaScript)
    vinyle = {
        "type": str(type_album or "").strip(),
        "annee": str(annee or "").strip(),
        "quantite": str(quantite or "1").strip(),
        "pays": str(pays or "").strip(),
        "genre": str(genre or "").strip(),
        "commentaire": str(commentaire or "").strip(),
        "artiste": str(artiste or "").strip(),
        "titre_face_a": str(titre_face_a or "").strip(),
        "titre_face_b": str(titre_face_b or "").strip()
    }
    liste_vinyles.append(vinyle)

print(f"🎉 Lignes chargées avec succès : {len(liste_vinyles)}")

# ---- CHARGEMENT DE LA LISTE D'ACHAT (WANTED) ----
print(f"\nAnalyse du fichier Wanted en cours ({EXCEL_ACHAT})...")
liste_wanted = []

if os.path.exists(EXCEL_ACHAT):
    wb_w = openpyxl.load_workbook(EXCEL_ACHAT, data_only=True)
    sheet_w = wb_w.active
    
    for row_idx in range(2, sheet_w.max_row + 1):
        artiste_titre = sheet_w.cell(row=row_idx, column=1).value  # Col A: Artistes / Titres
        note_c = sheet_w.cell(row=row_idx, column=3).value         # Col C: Note ("J'ai")
        url_image = sheet_w.cell(row=row_idx, column=6).value      # Col F: Lien image internet

        # FILTRE : Uniquement si la colonne C ("Note") est VIDE
        if artiste_titre and not str(note_c or "").strip():
            wanted_item = {
                "nom": str(artiste_titre).strip(),
                "image": str(url_image or "").strip()
            }
            liste_wanted.append(wanted_item)
else:
    print(f"Attention : Le fichier {EXCEL_ACHAT} n'a pas été trouvé.")

# ---- CONVERSION SÉCURISÉE EN JSON POUR ÉVITER TOUT PLANTAGE ----
# json.dumps met automatiquement des anti-slashs sur les apostrophes qui faisaient planter votre script !
json_vinyles = json.dumps(liste_vinyles, ensure_ascii=False)
json_wanted = json.dumps(liste_wanted, ensure_ascii=False)


# =========================================================================
# 3. CODE HTML COMPLET DE VOTRE PAGE PRINCIPALE (INDEX.HTML)
# =========================================================================
html_index_content = f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <title>Ma Collection de Vinyles</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            background-color: #f4f4f6;
            margin: 0;
            padding: 0;
        }}
        .header-bar {{
            background-color: #111;
            color: white;
            padding: 15px 30px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        .badge-total {{
            background-color: #ffcc00;
            color: #111;
            padding: 8px 12px;
            font-weight: bold;
            border-radius: 5px;
            font-size: 14px;
        }}
        .wanted-btn {{
            background-color: #111;
            color: #ffcc00;
            border: 2px solid #ffcc00;
            padding: 8px 15px;
            font-weight: bold;
            border-radius: 5px;
            cursor: pointer;
            text-decoration: none;
        }}
        .filter-section {{
            background-color: white;
            margin: 20px auto;
            padding: 20px;
            width: 90%;
            max-width: 1200px;
            border-radius: 8px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.05);
        }}
        .search-bar {{
            width: 100%;
            padding: 12px;
            font-size: 16px;
            border: 1px solid #ccc;
            border-radius: 5px;
            box-sizing: border-box;
            margin-bottom: 15px;
        }}
        .filter-group {{
            margin-bottom: 10px;
            display: flex;
            align-items: center;
            gap: 10px;
            flex-wrap: wrap;
        }}
        .filter-label {{
            font-weight: bold;
            width: 100px;
            text-transform: uppercase;
            font-size: 13px;
            color: #666;
        }}
        .btn-filter {{
            background-color: white;
            border: 1px solid #ccc;
            padding: 6px 12px;
            border-radius: 4px;
            cursor: pointer;
            font-weight: bold;
        }}
        .btn-filter.active {{
            background-color: #ffcc00;
            border-color: #ffcc00;
        }}
        .grid-container {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(180px, 1fr));
            gap: 20px;
            width: 90%;
            max-width: 1200px;
            margin: 20px auto;
        }}
        .card {{
            background-color: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 4px 6px rgba(0,0,0,0.05);
            display: flex;
            flex-direction: column;
            font-size: 12px;
            position: relative;
        }}
        .card-img {{
            width: 100%;
            height: 180px;
            object-fit: cover;
            background-color: #e0e0e0;
        }}
        .price-tag {{
            position: absolute;
            top: 10px;
            right: 10px;
            background-color: #ff3366;
            color: white;
            padding: 3px 6px;
            border-radius: 5px;
            font-weight: bold;
        }}
        .card-body {{
            padding: 10px;
            display: flex;
            flex-direction: column;
            gap: 4px;
        }}
        .tag-type {{
            background-color: #e6e6fa;
            color: #555;
            padding: 2px 4px;
            text-transform: uppercase;
            font-weight: bold;
            font-size: 10px;
            width: fit-content;
        }}
        .artiste-name {{
            font-weight: bold;
            font-size: 13px;
        }}
        .gray-line {{
            border-top: 1px solid #ddd;
            margin: 5px 0;
        }}
    </style>
</head>
<body>

    <div class="header-bar">
        <div class="badge-total">Total collection : {compteur_global}</div>
        <h2>Ma Collection de Vinyles</h2>
        <a href="Wanted.html" class="wanted-btn">Wanted ➔</a>
    </div>

    <div class="filter-section">
        <input type="text" id="searchInput" class="search-bar" placeholder="Rechercher un artiste, un titre, un pays...">
        
        <div class="filter-group">
            <span class="filter-label">Type :</span>
            <button class="btn-filter active" onclick="filterType('Tous', this)">Tous</button>
            <button class="btn-filter" onclick="filterType('ALBUM', this)">Album</button>
            <button class="btn-filter" onclick="filterType('COMPILS', this)">Compils</button>
            <button class="btn-filter" onclick="filterType('INTERVIEW', this)">Interview</button>
            <button class="btn-filter" onclick="filterType('JINGLE', this)">Jingle</button>
            <button class="btn-filter" onclick="filterType('MEDLEY', this)">Medley</button>
            <button class="btn-filter" onclick="filterType('SINGLE', this)">Single</button>
        </div>

        <div class="filter-group">
            <span class="filter-label">Genre :</span>
            <select id="genreSelect" onchange="applyFilters()" style="padding: 6px; border-radius: 4px;">
                <option value="Tous">Tous les genres</option>
            </select>
        </div>

        <div style="margin-top: 15px; font-weight: bold; color: #444;">
            Disques affichés : <span id="displayedCount">0</span>
        </div>
    </div>

    <div class="grid-container" id="gridContainer"></div>

    <script>
        // Injection de la liste brute nettoyée
        const dataVinyles = {json_vinyles};
        
        let currentType = 'Tous';
        
        // Remplir la liste déroulante des genres dynamiquement
        const genreSelect = document.getElementById('genreSelect');
        const genresUniques = [...new Set(dataVinyles.map(v => v.genre).filter(g => g))].sort();
        genresUniques.forEach(g => {{
            const opt = document.createElement('option');
            opt.value = g;
            opt.innerText = g;
            genreSelect.appendChild(opt);
        }});

        function filterType(type, element) {{
            document.querySelectorAll('.filter-group button').forEach(b => b.classList.remove('active'));
            element.classList.add('active');
            currentType = type;
            applyFilters();
        }}

        function applyFilters() {{
            const searchVal = document.getElementById('searchInput').value.toLowerCase();
            const selectedGenre = document.getElementById('genreSelect').value;
            
            const filtered = dataVinyles.filter(v => {{
                const matchesSearch = v.artiste.toLowerCase().includes(searchVal) || 
                                      v.titre_face_a.toLowerCase().includes(searchVal) || 
                                      v.pays.toLowerCase().includes(searchVal);
                const matchesType = (currentType === 'Tous' || v.type.toUpperCase() === currentType);
                const matchesGenre = (selectedGenre === 'Tous' || v.genre === selectedGenre);
                
                return matchesSearch && matchesType && matchesGenre;
            }});

            renderGrid(filtered);
        }}

        function renderGrid(items) {{
            const container = document.getElementById('gridContainer');
            container.innerHTML = '';
            document.getElementById('displayedCount').innerText = items.length;

            items.forEach(item => {{
                // Gestion du prix ou de la mention spéciale dans la Col H (Commentaire)
                let priceHtml = '';
                if(item.commentaire && item.commentaire.includes('€')) {{
                    priceHtml = `<div class="price-tag">${{item.commentaire}}</div>`;
                }} else if (item.commentaire === 'PINTO') {{
                    priceHtml = `<div class="price-tag" style="background:#ffcc00; color:black;">PINTO</div>`;
                }}

                // Construction dynamique de la pochette
                const card = document.createElement('div');
                card.className = 'card';
                card.innerHTML = `
                    ${{priceHtml}}
                    <div class="card-img" style="display:flex; align-items:center; justify-content:center; font-weight:bold; color:#aaa; background:#222; height:180px;">🎵 VINYLE</div>
                    <div class="card-body">
                        <div class="tag-type">${{item.type}}</div>
                        <div class="artiste-name">${{item.artiste}}</div>
                        <div style="color:#555;"><b>Genre :</b> ${{item.genre}}</div>
                        <div style="color:#555;"><b>Année :</b> ${{item.annee}} | <b>Pays :</b> ${{item.pays}}</div>
                        <div class="gray-line"></div>
                        <div style="font-size:11px; color:#111;"><b>Face A :</b> ${{item.titre_face_a || 'Non renseigné'}}</div>
                        <div style="font-size:11px; color:#111; margin-top:2px;"><b>Face B :</b> ${{item.titre_face_b || 'Non renseigné'}}</div>
                    </div>
                `;
                container.appendChild(card);
            }});
        }}

        // Écouteur sur la recherche texte
        document.getElementById('searchInput').addEventListener('input', applyFilters);

        // Premier affichage au chargement
        applyFilters();
    </script>
</body>
</html>
"""

with open(INDEX_HTML, "w", encoding="utf-8") as f:
    f.write(html_index_content)
print(f"💾 Fichier {INDEX_HTML} généré avec succès !")


# =========================================================================
# 4. CODE HTML COMPLET DE LA PAGE WANTED (WANTED.HTML)
# =========================================================================
html_wanted_content = f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <title>Ma Liste de Recherche (Wanted)</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            background-color: #f4f4f6;
            margin: 0;
            padding: 0;
        }}
        .header-bar {{
            background-color: #111;
            color: white;
            padding: 15px 30px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        .collection-btn {{
            background-color: white;
            color: #111;
            padding: 8px 15px;
            font-weight: bold;
            border-radius: 5px;
            text-decoration: none;
        }}
        .wanted-container {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(220px, 1fr));
            gap: 20px;
            width: 90%;
            max-width: 1200px;
            margin: 30px auto;
        }}
        .wanted-card {{
            background-color: white;
            border-radius: 8px;
            padding: 15px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.05);
            text-align: center;
        }}
        .wanted-img {{
            width: 100%;
            height: 200px;
            object-fit: cover;
            border-radius: 5px;
            margin-top: 10px;
            background-color: #eee;
        }}
    </style>
</head>
<body>

    <div class="header-bar">
        <a href="index.html" class="collection-btn">📁 Collection</a>
        <h2>Ma Liste de Recherche (Wanted)</h2>
        <div style="font-weight: bold; color: #ffcc00;">Disques recherchés : <span id="wanted-count">0</span></div>
    </div>

    <div class="wanted-container" id="wantedContainer"></div>

    <script>
        const dataWanted = {json_wanted};
        document.getElementById('wanted-count').innerText = dataWanted.length;
        
        const container = document.getElementById('wantedContainer');
        dataWanted.forEach(item => {{
            const card = document.createElement('div');
            card.className = 'wanted-card';
            
            // On affiche le nom (Artiste - Titre) et l'image récupérée depuis la colonne F
            card.innerHTML = `
                <div style="font-weight: bold; font-size: 14px; min-height: 40px; color:#222;">${{item.nom}}</div>
                ${{item.image ? `<img src="${{item.image}}" class="wanted-img" alt="Pochette">` : '<div style="height:200px; background:#ddd; margin-top:10px; display:flex; align-items:center; justify-content:center; color:#888; border-radius:5px;">Pas d\'image</div>'}}
            `;
            container.appendChild(card);
        }});
    </script>
</body>
</html>
"""

with open(WANTED_HTML, "w", encoding="utf-8") as f:
    f.write(html_wanted_content)
print(f"💾 Fichier {WANTED_HTML} généré avec succès ! Total recherchés : {len(liste_wanted)}")

print("\n2. Envoi des mises à jour sur GitHub...")
os.system("git add index.html Wanted.html")
os.system('git commit -m "Correction complete de l affichage index et wanted"')
os.system("git push origin main")
print("\n🟢 Tout est réparé et synchronisé ! Vous pouvez rafraîchir votre navigateur.")