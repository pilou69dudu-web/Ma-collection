import json
import os
import re
import shutil
import subprocess
import sys
import time
import urllib.request


def installer_si_manquant(module_name):
  try:
    __import__(module_name)
  except ImportError:
    print(f"📦 Installation de {module_name}...")
    subprocess.check_call(
        [sys.executable, "-m", "pip", "install", module_name]
    )


installer_si_manquant("openpyxl")
installer_si_manquant("cloudscraper")
installer_si_manquant("beautifulsoup4")

from bs4 import BeautifulSoup
import cloudscraper
import openpyxl

# =====================================================================
# CONFIGURATION
# =====================================================================
FICHIER_ORIGINAL = "00_Mes vinyles.xlsx"
FICHIER_TMP = "00_Mes vinyles_TEMP.xlsx"

# Scraper simulant Chrome sous Windows
scraper = cloudscraper.create_scraper(
    browser={"browser": "chrome", "platform": "windows", "desktop": True}
)


# =====================================================================
# RECUPERATION DU TAUX DE CHANGE EN DIRECT
# =====================================================================
def obtenir_taux_usd_eur():
  """Récupère le taux de change USD -> EUR en direct via une API gratuite.

  Utilise 0.86 en taux de secours si l'API est inaccessible.
  """
  url_api = "https://api.exchangerate-api.com/v4/latest/USD"
  try:
    req = urllib.request.Request(
        url_api, headers={"User-Agent": "Mozilla/5.0"}
    )
    with urllib.request.urlopen(req, timeout=5) as response:
      data = json.loads(response.read().decode())
      taux = data.get("rates", {}).get("EUR")
      if taux:
        print(f"💱 Taux de change du jour récupéré : 1 USD = {taux} EUR")
        return float(taux)
  except Exception as e:
    print(
        f"⚠️ Impossible de récupérer le taux en direct ({e}). Utilisation du"
        " taux par défaut : 0.86"
    )

  return 0.86


TAUX_USD_EUR = obtenir_taux_usd_eur()


# =====================================================================


def extraire_id_discogs(url):
  if not url:
    return None
  match = re.search(r"release/(\d+)", str(url))
  return match.group(1) if match else None


def recuperer_prix_high(lien_excel, release_id):
  """Récupère la cote 'Élevée' / 'High' depuis Discogs.

  Combine API stats, JSON interne et parsing direct du texte HTML.
  """
  rid = release_id or extraire_id_discogs(lien_excel)
  if not rid:
    return None

  headers_api = {
      "User-Agent": (
          "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
          " (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
      ),
      "Referer": f"https://www.discogs.com/fr/release/{rid}",
      "Accept": "application/json, text/plain, */*",
  }

  # -----------------------------------------------------------------
  # METHODE 1 : Endpoint API interne de statistiques Marketplace
  # -----------------------------------------------------------------
  url_stats = f"https://www.discogs.com/marketplace/stats/{rid}"
  try:
    res = scraper.get(url_stats, headers=headers_api, timeout=8)
    if res.status_code == 200:
      data = res.json()
      high_val = None
      if "high" in data:
        high_val = (
            data["high"].get("value")
            if isinstance(data["high"], dict)
            else data["high"]
        )
      elif "highest_price" in data:
        high_val = (
            data["highest_price"].get("value")
            if isinstance(data["highest_price"], dict)
            else data["highest_price"]
        )
      elif "price" in data and "high" in data["price"]:
        high_val = data["price"]["high"]

      if high_val and float(high_val) > 0:
        prix_usd = float(high_val)
        prix_eur = round(prix_usd * TAUX_USD_EUR)
        return f"{prix_eur} €"
  except Exception:
    pass

  # -----------------------------------------------------------------
  # METHODE 2 & 3 : Parsing de la page HTML
  # -----------------------------------------------------------------
  url_page = f"https://www.discogs.com/fr/release/{rid}"
  try:
    res_page = scraper.get(url_page, headers=headers_api, timeout=10)
    if res_page.status_code == 200:
      html_text = res_page.text

      # METHODE 2 : JSON-LD / __NEXT_DATA__ embarqué dans le code source
      soup = BeautifulSoup(html_text, "html.parser")
      for script in soup.find_all("script"):
        stext = script.string or ""
        if "high" in stext or "highest" in stext or "last_sold" in stext:
          matches = re.findall(
              r'"high"\s*:\s*(?:{\s*"value"\s*:\s*|)"?(\d+[\.,]?\d*)"?', stext
          )
          for m in matches:
            v = float(m.replace(",", "."))
            if v > 0:
              prix_eur = round(v * TAUX_USD_EUR)
              return f"{prix_eur} €"

      matches_highprice = re.findall(
          r'"highPrice"\s*:\s*"?(\d+[\.,]?\d*)"?', html_text
      )
      if matches_highprice:
        v = float(matches_highprice[0].replace(",", "."))
        prix_eur = round(v * TAUX_USD_EUR)
        return f"{prix_eur} €"

      # METHODE 3 : Extraction directe du texte "Élevée:" dans le HTML
      # Gère la mise en page visible comme sur la capture : "Élevée: 20,85 €"
      match_textuel = re.search(
          r"(?:Élevée|High)\s*:\s*([\d\s\.,]+)\s*(?:€|\$|EUR|USD)?",
          html_text,
          re.IGNORECASE,
      )
      if match_textuel:
        raw_val = (
            match_textuel.group(1)
            .replace("\xa0", "")
            .replace(" ", "")
            .replace(",", ".")
        )
        val = float(raw_val)
        if val > 0:
          # Si le montant extrait est déjà affiché en EUR sur la page FR, on arrondit directement
          if "€" in match_textuel.group(0) or "EUR" in match_textuel.group(0):
            return f"{int(round(val))} €"
          else:
            prix_eur = round(val * TAUX_USD_EUR)
            return f"{prix_eur} €"

  except Exception:
    pass

  return None


def main():
  print(
      "====================================================================="
  )
  print(
      "   MISE À JOUR DES PRIX HAUTS ('ÉLEVÉE') VIA DISCOGS STATS (USD ->"
      " EUR)  "
  )
  print(
      "====================================================================="
  )

  if not os.path.exists(FICHIER_ORIGINAL):
    print(f"❌ Fichier non trouvé : {FICHIER_ORIGINAL}")
    return

  if os.path.exists(FICHIER_TMP):
    try:
      os.remove(FICHIER_TMP)
    except Exception:
      print("❌ Fermez Excel avant de démarrer.")
      return

  shutil.copyfile(FICHIER_ORIGINAL, FICHIER_TMP)
  print(f"💾 Copie de travail créée : {FICHIER_TMP}\n")

  wb = openpyxl.load_workbook(FICHIER_TMP)
  ws = wb.active

  # Détection automatique de l'en-tête et des colonnes
  header_row = 1
  for r in range(1, 4):
    cells = [
        str(ws.cell(row=r, column=c).value).lower()
        for c in range(1, ws.max_column + 1)
    ]
    if "artiste" in cells or "lien" in cells:
      header_row = r
      break

  headers = {
      str(ws.cell(row=header_row, column=c).value).strip(): c
      for c in range(1, ws.max_column + 1)
  }
  col_lien = headers.get("Lien", 17)
  col_artiste = headers.get("ARTISTE", 9)
  col_prix = headers.get("Prix Haut", 18)

  total = ws.max_row - header_row

  for row_idx in range(header_row + 1, ws.max_row + 1):
    artiste = (
        ws.cell(row=row_idx, column=col_artiste).value or f"Ligne {row_idx}"
    )
    lien = ws.cell(row=row_idx, column=col_lien).value

    release_id = extraire_id_discogs(lien)

    if not lien and not release_id:
      continue

    print(
        f"🔍 [{row_idx - header_row}/{total}] {artiste} (ID: {release_id})... ",
        end="",
        flush=True,
    )

    prix_high = recuperer_prix_high(lien, release_id)

    if prix_high:
      ws.cell(row=row_idx, column=col_prix).value = prix_high
      print(f"✅ ÉLEVÉE : {prix_high}")
    else:
      print("💾 (Pas de cote / Non vendu)")

    # Sauvegarde toutes les 5 lignes
    if (row_idx - header_row) % 5 == 0:
      wb.save(FICHIER_TMP)

    time.sleep(1.2)

  wb.save(FICHIER_TMP)
  wb.close()

  try:
    shutil.move(FICHIER_TMP, FICHIER_ORIGINAL)
    print("\n✅ Terminé ! Votre fichier Excel est entièrement mis à jour.")
  except Exception:
    print(
        f"\n⚠️ Sauvegardé dans '{FICHIER_TMP}'. Fermez Excel pour valider le"
        " remplacement."
    )


if __name__ == "__main__":
  main()